import os
import sc2reader
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

FILE_NAME = "sc2.xlsx"
FOLDER = 'replays_/'
MINIMUM_PLAYERS = 2
MAXIMUM_AI = 1
SUM_STRING = '=COUNTIFS(Players!$C:$C,$A{},Players!${}:${},"{}")'

class Game():
    def __init__(self, replay):
        self.id = f'{replay.date.date()}-{replay.raw_data["replay.initData.backup"]["game_description"]["random_value"]}'
        self.map = replay.map
        self.file_name = replay.filename
        self.type = replay.type
        self.datetime = replay.date
        self.category = replay.category
        self.length = replay.length.seconds
        self.build = replay.build
        self.release_string = replay.release_string
        self.teams = [team for team in replay.teams]
        self.matchup = "v".join(sorted(team.lineup for team in self.teams))
        self.players = sum((team.players for team in self.teams), [])

def fixColumnWidth(ws):
    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

    ws.column_dimensions = dim_holder

def clean_name(name):
    if name.startswith("A.I."):
        return "A.I. " + name[name.index('('):]
    else:
        return name

replays = sc2reader.load_replays(FOLDER, load_level=2, load_map=True)

wb = Workbook()

overviewWS = wb.active
overviewWS.title = "Overview"
overviewWS.append(['Name', 'Win', 'Loss', 'Unknown', 'Ratio', 'Terran', 'Zerg', 'Protoss'])

gamesWS = wb.create_sheet("Games")
gamesWS.append(['Date', 'Map', 'Type', 'Length', 'Team 1', 'Team 2'])

playersWS = wb.create_sheet("Players")
playersWS.append(['Date', 'Map', 'Name', 'Race', 'Result', 'Handicap'])

all_players = []
games = {}

for replay in replays:
    game = Game(replay)

    if len(game.players) <= MINIMUM_PLAYERS:
        continue

    if len([p for p in game.players if p.is_human == False]) > MAXIMUM_AI:
        continue

    if game.id in games:
        if (games[game.id].teams[0].result == None):
            games[game.id] = game
    else:
        games[game.id] = game

    print(f'{game.datetime} - {game.id} - {game.map.name} {game.type}, Game length: {round(game.length / 60)}:{game.length % 60}')


for id, game in games.items():
    game_result = []

    for team in game.teams:
        players = ", ".join([f'{clean_name(player.name)} ({player.play_race})' for player in team.players])
        game_result.append(f'{team.number} - {team.result} - {players}')

        for player in team.players:
            playersWS.append([game.datetime, game.map.name, clean_name(player.name), player.play_race, team.result, player.handicap])
            if (clean_name(player.name) not in all_players):
                all_players.append(clean_name(player.name))

    gamesWS.append([game.datetime, game.map.name, game.type, f'{round(game.length / 60)}:{game.length % 60}'] + game_result)


for index, player in enumerate(sorted(all_players, key=str.lower, reverse=True)):
    index += 2
    overviewWS.append([
        player, 
        SUM_STRING.format(index, 'E', 'E', 'Win'), 
        SUM_STRING.format(index, 'E', 'E', 'Loss'),
        SUM_STRING.format(index, 'E', 'E', ''),
        '=IF($C{i}>0,$B{i}/$C{i},0)'.format(i=index),
        SUM_STRING.format(index, 'D', 'D', 'Terran'), 
        SUM_STRING.format(index, 'D', 'D', 'Zerg'), 
        SUM_STRING.format(index, 'D', 'D', 'Protoss')
    ])

overviewWS.column_dimensions['A'].width = 17

gamesWS.column_dimensions['A'].width = 18
gamesWS.column_dimensions['B'].width = 18
gamesWS.column_dimensions['E'].width = 60
gamesWS.column_dimensions['F'].width = 60

playersWS.column_dimensions['A'].width = 18
playersWS.column_dimensions['B'].width = 18
playersWS.column_dimensions['C'].width = 18

overviewWS.freeze_panes = overviewWS['A2']
gamesWS.freeze_panes = gamesWS['A2']
playersWS.freeze_panes = playersWS['A2']

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Games"

data = Reference(overviewWS, min_col=2, min_row=1, max_row=len(all_players) + 1, max_col=4)
cats = Reference(overviewWS, min_col=1, min_row=2, max_row=len(all_players) + 1)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
overviewWS.add_chart(chart1, "I2")

chart2 = BarChart()
chart2.type = "col"
chart2.style = 10
chart2.title = "Race"

data = Reference(overviewWS, min_col=6, min_row=1, max_row=len(all_players) + 1, max_col=8)
cats = Reference(overviewWS, min_col=1, min_row=2, max_row=len(all_players) + 1)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
chart2.shape = 4
overviewWS.add_chart(chart2, "I18")

# Save the file
wb.save(FILE_NAME)